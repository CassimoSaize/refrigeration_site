from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .forms import ContactForm

from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from .models import OrcamentoSolicitacao
from .forms import OrcamentoForm, OrcamentoStatusForm


from .models import HeroSection, Servico, Sobre, ImagemGaleria, ContatoInfo, RedeSocial
from .forms import (HeroForm, ServicoForm, SobreForm, ImagemGaleriaForm,
                    ContatoInfoForm, RedeSocialForm)


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
import tempfile


from xhtml2pdf import pisa
from django.template.loader import render_to_string
from io import BytesIO



def home(request):
    
    # Processar o formulário se for POST
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mensagem enviada com sucesso! Entrarei em contato em breve.')
            return redirect('home')  # redireciona para a mesma página (evita reenvio)
        else:
            messages.error(request, 'Erro ao enviar. Verifique os campos.')
    else:
        form = ContactForm()  # formulário vazio para GET
        
    # Cada modelo com poucos registros (geralmente um único)
    hero = HeroSection.objects.first()
    sobre = Sobre.objects.first()
    contato = ContatoInfo.objects.first()

    # Listas (vários registros)
    servicos = Servico.objects.all()
    galeria = ImagemGaleria.objects.all()
    redes_sociais = RedeSocial.objects.all()

    context = {
        'hero': hero,
        'sobre': sobre,
        'contato': contato,
        'servicos': servicos,
        'galeria': galeria,
        'redes_sociais': redes_sociais,
        'form': form,  # ← ESSA LINHA É ESSENCIAL
    }
    return render(request, 'index.html', context)


#---------------------------------------------------------------------------------------------
#---------------------------------------------------------------------------------------------




@login_required
def dashboard_index(request):
    """Página principal do dashboard com resumo e links."""
    context = {
        'total_servicos': Servico.objects.count(),
        'total_galeria': ImagemGaleria.objects.count(),
        'total_redes': RedeSocial.objects.count(),
    }
    return render(request, 'dashboard/index.html', context)

# ---------- Hero ----------
def hero_list(request):
    hero = HeroSection.objects.first()

    context = {
        'hero': hero
    }

    return render(request, 'index.html', context)

@login_required
def hero_edit(request):
    hero = HeroSection.objects.first()
    if not hero:
        hero = HeroSection.objects.create()
    if request.method == 'POST':
        form = HeroForm(request.POST, instance=hero)
        if form.is_valid():
            form.save()
            messages.success(request, 'Hero atualizado com sucesso!')
            return redirect('hero_edit')
    else:
        form = HeroForm(instance=hero)
    return render(request, 'dashboard/hero_form.html', {'form': form})

# ---------- Serviços (CRUD) ----------
@login_required
def servico_list(request):
    servicos = Servico.objects.all()
    return render(request, 'dashboard/servico_list.html', {'servicos': servicos})

@login_required
def servico_create(request):
    if request.method == 'POST':
        form = ServicoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço criado!')
            return redirect('servico_list')
    else:
        form = ServicoForm()
    return render(request, 'dashboard/servico_form.html', {'form': form})

@login_required
def servico_edit(request, pk):
    servico = get_object_or_404(Servico, pk=pk)
    if request.method == 'POST':
        form = ServicoForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço atualizado!')
            return redirect('servico_list')
    else:
        form = ServicoForm(instance=servico)
    return render(request, 'dashboard/servico_form.html', {'form': form})

@login_required
def servico_delete(request, pk):
    servico = get_object_or_404(Servico, pk=pk)
    if request.method == 'POST':
        servico.delete()
        messages.success(request, 'Serviço removido!')
        return redirect('servico_list')
    return render(request, 'dashboard/confirm_delete.html', {'object': servico})

# Views similares para ImagemGaleria, Sobre, ContatoInfo, RedeSocial...
# Vou resumir as mais importantes.

@login_required
def galeria_list(request):
    imagens = ImagemGaleria.objects.all()
    return render(request, 'dashboard/galeria_list.html', {'imagens': imagens})

@login_required
def galeria_create(request):
    if request.method == 'POST':
        form = ImagemGaleriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Imagem adicionada!')
            return redirect('galeria_list')
    else:
        form = ImagemGaleriaForm()
    return render(request, 'dashboard/galeria_form.html', {'form': form})

@login_required
def galeria_edit(request, pk):
    imagem = get_object_or_404(ImagemGaleria, pk=pk)
    if request.method == 'POST':
        form = ImagemGaleriaForm(request.POST, request.FILES, instance=imagem)
        if form.is_valid():
            form.save()
            messages.success(request, 'Imagem atualizada!')
            return redirect('dashboard:galeria_list')
    else:
        form = ImagemGaleriaForm(instance=imagem)
    return render(request, 'dashboard/galeria_form.html', {'form': form})

@login_required
def galeria_delete(request, pk):
    imagem = get_object_or_404(ImagemGaleria, pk=pk)
    if request.method == 'POST':
        imagem.delete()
        messages.success(request, 'Imagem excluída!')
        return redirect('galeria_list')
    return render(request, 'dashboard/confirm_delete.html', {'object': imagem})

@login_required
def sobre_edit(request):
    sobre = Sobre.objects.first()
    if not sobre:
        sobre = Sobre.objects.create(texto='', diferenciais='')
    if request.method == 'POST':
        form = SobreForm(request.POST, instance=sobre)
        if form.is_valid():
            form.save()
            messages.success(request, 'Seção Sobre atualizada!')
            return redirect('sobre_edit')
    else:
        form = SobreForm(instance=sobre)
    return render(request, 'dashboard/sobre_form.html', {'form': form})



def contato_create(request):
    
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mensagem enviada com sucesso! Entrarei em contato em breve.')
            return redirect('home')
        else:
            messages.error(request, 'Erro ao enviar. Verifique os campos.')
    else:
        form = ContactForm()
    
    return render(request, 'index.html', {'form': form})


@login_required
def contato_edit(request):
    contato = ContatoInfo.objects.first()
    if not contato:
        contato = ContatoInfo.objects.create()
    if request.method == 'POST':
        form = ContatoInfoForm(request.POST, instance=contato)
        if form.is_valid():
            form.save()
            messages.success(request, 'Informações de contato atualizadas!')
            return redirect('contato_edit')
    else:
        form = ContatoInfoForm(instance=contato)
    return render(request, 'dashboard/contato_form.html', {'form': form})


#----------------------------------------------------------------------------------
#----------------------------------------------------------------------------------



def solicitar_orcamento(request):
    """View pública para solicitar orçamento"""
    if request.method == 'POST':
        form = OrcamentoForm(request.POST)
        if form.is_valid():
            orcamento = form.save()

            # Enviar email de confirmação para o cliente
            try:
                send_mail(
                    subject=f'Confirmação de Solicitação de Orçamento - Cassimo Refrigeração',
                    message=f"""
                    Olá {orcamento.nome}!

                    Recebemos sua solicitação de orçamento para {orcamento.get_tipo_servico_display()}.

                    Detalhes da solicitação:
                    - Serviço: {orcamento.get_tipo_servico_display()}
                    - Equipamento: {orcamento.get_tipo_equipamento_display() or 'Não especificado'}
                    - Quantidade: {orcamento.quantidade}

                    Entraremos em contato em breve através do telefone {orcamento.telefone} ou e-mail {orcamento.email}.

                    Protocolo: #{orcamento.id}
                    Data: {orcamento.created_at.strftime('%d/%m/%Y %H:%M')}

                    Atenciosamente,
                    Equipe Cassimo Refrigeração & Climatização
                    """,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[orcamento.email],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Erro ao enviar email: {e}")

            # Enviar notificação para o admin (pode ser via WhatsApp ou email)
            try:
                send_mail(
                    subject=f'Nova Solicitação de Orçamento #{orcamento.id}',
                    message=f"""
                    Nova solicitação de orçamento recebida!

                    Cliente: {orcamento.nome}
                    Telefone: {orcamento.telefone}
                    Serviço: {orcamento.get_tipo_servico_display()}

                    Acesse o dashboard para visualizar:
                    {request.build_absolute_uri('/dashboard/orcamentos/')}
                    """,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[settings.ADMIN_EMAIL],
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Erro ao enviar notificação: {e}")

            messages.success(request, 'Solicitação enviada com sucesso! Entraremos em contato em breve.')
            return redirect('solicitar_orcamento')
    else:
        form = OrcamentoForm()

    return render(request, 'dashboard/orcamento_form.html', {'form': form})

# Dashboard Views
@login_required
def orcamento_list(request):
    """Lista todas as solicitações de orçamento"""
    orcamentos = OrcamentoSolicitacao.objects.all()

    # Filtros
    status_filter = request.GET.get('status')
    if status_filter:
        orcamentos = orcamentos.filter(status=status_filter)

    tipo_filter = request.GET.get('tipo')
    if tipo_filter:
        orcamentos = orcamentos.filter(tipo_servico=tipo_filter)

    # Estatísticas
    stats = {
        'total': OrcamentoSolicitacao.objects.count(),
        'pending': OrcamentoSolicitacao.objects.filter(status='pending').count(),
        'viewed': OrcamentoSolicitacao.objects.filter(status='viewed').count(),
        'contacted': OrcamentoSolicitacao.objects.filter(status='contacted').count(),
        'budget_sent': OrcamentoSolicitacao.objects.filter(status='budget_sent').count(),
        'approved': OrcamentoSolicitacao.objects.filter(status='approved').count(),
        'completed': OrcamentoSolicitacao.objects.filter(status='completed').count(),
    }

    return render(request, 'dashboard/orcamento_list.html', {
        'orcamentos': orcamentos,
        'stats': stats,
        'current_status': status_filter,
        'current_tipo': tipo_filter,
    })

@login_required
def orcamento_detail(request, pk):
    """Detalhes de uma solicitação de orçamento"""
    orcamento = get_object_or_404(OrcamentoSolicitacao, pk=pk)

    # Marcar como visualizado
    if orcamento.status == 'pending':
        orcamento.mark_as_viewed()

    if request.method == 'POST':
        form = OrcamentoStatusForm(request.POST, instance=orcamento)
        if form.is_valid():
            form.save()
            messages.success(request, f'Orçamento #{orcamento.id} atualizado com sucesso!')
            return redirect('orcamento_detail', pk=orcamento.pk)
    else:
        form = OrcamentoStatusForm(instance=orcamento)

    return render(request, 'dashboard/orcamento_detail.html', {
        'orcamento': orcamento,
        'form': form,
    })

@login_required
def orcamento_delete(request, pk):
    """Excluir solicitação de orçamento"""
    orcamento = get_object_or_404(OrcamentoSolicitacao, pk=pk)
    if request.method == 'POST':
        orcamento.delete()
        messages.success(request, 'Solicitação excluída com sucesso!')
        return redirect('orcamento_list')
    return render(request, 'dashboard/confirm_delete.html', {'object': orcamento})


#----------------------------------------------------------------------------------
#----------------------------------------------------------------------------------

@login_required
def orcamento_export_pdf(request, pk):
    """Exporta o orçamento em PDF usando xhtml2pdf (sem dependências externas)"""
    orcamento = get_object_or_404(OrcamentoSolicitacao, pk=pk)
    
    # Renderiza o template HTML
    html_string = render_to_string('dashboard/orcamento_pdf.html', {'orcamento': orcamento})
    
    # Cria um objeto BytesIO para o PDF
    result = BytesIO()
    
    # Converte HTML para PDF
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="orcamento_{orcamento.id}.pdf"'
        response.write(result.getvalue())
        return response
    else:
        return HttpResponse('Erro ao gerar PDF: %s' % pdf.err, status=500)

@login_required
def orcamento_export_excel(request, pk):
    """Exporta o orçamento em Excel (.xlsx) com formatação profissional"""
    orcamento = get_object_or_404(OrcamentoSolicitacao, pk=pk)
    
    # Cria workbook e seleciona a planilha ativa
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="007BFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    # Dados do orçamento em pares (rótulo, valor)
    dados = [
        ("Nº do Orçamento", orcamento.id),
        ("Data da Solicitação", orcamento.created_at.strftime("%d/%m/%Y %H:%M")),
        ("Última Atualização", orcamento.updated_at.strftime("%d/%m/%Y %H:%M")),
        ("Status", orcamento.get_status_display()),
        ("--- DADOS DO CLIENTE ---", ""),
        ("Nome", orcamento.nome),
        ("E-mail", orcamento.email),
        ("Telefone", orcamento.telefone),
        ("Endereço", orcamento.endereco or "Não informado"),
        ("--- DETALHES DO SERVIÇO ---", ""),
        ("Tipo de serviço", orcamento.get_tipo_servico_display()),
        ("Equipamento", orcamento.get_tipo_equipamento_display() or "Não especificado"),
        ("Quantidade", f"{orcamento.quantidade} unidade(s)"),
        ("Descrição", orcamento.descricao),
        ("--- PREFERÊNCIAS ---", ""),
        ("Data preferencial", orcamento.data_preferencial.strftime("%d/%m/%Y") if orcamento.data_preferencial else "Não informada"),
        ("Horário preferencial", orcamento.get_horario_preferencial_display() or "Não informado"),
        ("--- INFORMAÇÕES FINANCEIRAS ---", ""),
        ("Valor orçado", f"{orcamento.valor_orcamento:,.2f} MZN" if orcamento.valor_orcamento else "Não definido"),
        ("Observações internas", orcamento.observacoes or "Nenhuma observação"),
    ]
    
    # Preenche a planilha
    for i, (label, value) in enumerate(dados, start=1):
        # Célula do rótulo
        cell_label = ws.cell(row=i, column=1, value=label)
        cell_label.font = header_font if "---" in label else Font(bold=True)
        cell_label.fill = header_fill if "---" in label else PatternFill("solid", fgColor="F0F0F0")
        cell_label.alignment = header_alignment if "---" in label else Alignment(horizontal="left", vertical="center")
        cell_label.border = border
        
        # Célula do valor
        cell_value = ws.cell(row=i, column=2, value=value)
        cell_value.alignment = cell_alignment
        cell_value.border = border
        if label == "Valor orçado" and orcamento.valor_orcamento:
            cell_value.font = Font(bold=True, color="28A745")
    
    # Ajuste especial para a descrição (pode conter múltiplas linhas)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "Descrição" in str(cell.value):
                ws.row_dimensions[cell.row].height = 60
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Valor da descrição está na coluna B da mesma linha
                desc_cell = ws.cell(row=cell.row, column=2)
                desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Adiciona uma segunda aba com histórico de alterações (opcional)
    if hasattr(orcamento, 'logs') and orcamento.logs.exists():
        ws2 = wb.create_sheet("Histórico")
        ws2.append(["Data/Hora", "Usuário", "Ação", "Detalhes"])
        for log in orcamento.logs.all().order_by('-created_at'):
            ws2.append([
                log.created_at.strftime("%d/%m/%Y %H:%M"),
                log.user.username if log.user else "Sistema",
                log.action,
                log.details
            ])
    
    # Prepara a resposta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="orcamento_{orcamento.id}.xlsx"'
    wb.save(response)
    return response

